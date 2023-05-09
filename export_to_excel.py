import os
import openpyxl
import datetime
import time
from pathlib import Path
import re

MAX_RECORD_NUMBER_PER_FILE = 10000


def set_sheet_header(sheet):
    sheet['A1'] = '논문번호'
    sheet['B1'] = '참고문헌'
    sheet['C1'] = 'DOI(url)'
    sheet['D1'] = 'DOI2(url)'
    sheet['E1'] = 'Pubmed(url)'
    sheet['F1'] = 'Pubmed Centrial(url)'


def export_to_excel(target_dir):
    """지정된 폴더의 파일을 읽어서 엑셀에 도출."""
    try:
        excel_path = Path(target_dir) / 'excel'
        os.makedirs(excel_path, exist_ok=True)  # 결과 폴더 생성

        wb = openpyxl.Workbook()  # 빈 엑셀 파일 만들기
        sheet = wb.active
        set_sheet_header(sheet)

        for filename in os.listdir(target_dir):
            file_path = Path(target_dir) / filename
            if os.path.isfile(file_path) and filename.endswith('.txt'):
                artical_no = filename[0:-4]  # 논문번호

                with open(file_path, 'r', encoding='utf8') as f:
                    content = f.read()  # 결과 파일 내용 읽기
                    result_item_list = content.strip().split('\n\n')  # 결과 항목 리스트

                    if (sheet.max_row + len(result_item_list)) > MAX_RECORD_NUMBER_PER_FILE:
                        now = datetime.datetime.now()
                        wb.save(excel_path /
                                f"참고문헌_일괄등록_{now.strftime('%Y%m%d_%H%M%S')}.xlsx")
                        time.sleep(1)

                        wb = openpyxl.Workbook()  # 빈 엑셀 파일 만들기
                        sheet = wb.active
                        set_sheet_header(sheet)

                    for result_item in result_item_list:
                        detail_list = result_item.split('\n')  # 단건 결과 항목 상세
                        # print(detail_list)
                        if len(detail_list) > 0:
                            row_num = sheet.max_row + 1  # 현재 row번호
                            # print(row_num)
                            sheet.cell(row=row_num,
                                       column=1).value = artical_no  # 논문번호 설정
                            sheet.cell(row=row_num,
                                       column=2).value = detail_list[0]  # 참고문헌 설정
                            if len(detail_list) >= 2:  # 여기서 부터 DOI url / PMid(PMCid)에 대해서 존재 여부 판단
                                doi_url_list = list(filter(
                                    lambda ele: ele.startswith('https://doi.org'), detail_list[1:]))  # DOI (url) 리스트 가져오기
                                if len(doi_url_list) > 0:
                                    sheet.cell(
                                        row=row_num, column=3).value = doi_url_list[0]  # DOI (url) 설정
                                if len(doi_url_list) > 1:
                                    sheet.cell(
                                        row=row_num, column=4).value = doi_url_list[1]  # DOI2 (url) 설정

                                pm_reg = r'PMC?id:(\d)+'
                                pmid_exists = len(list(filter(  # PMid 혹은 PMCid가 존재하는지 판단
                                    lambda ele: re.compile(pm_reg).search(ele), detail_list[1:]))) > 0

                                if pmid_exists:
                                    pmid_record = detail_list[1:][-1]
                                    pm_id_mo = re.compile(
                                        r'PMid:(\d+)').search(pmid_record)
                                    if pm_id_mo:
                                        pm_id = pm_id_mo.group(1)
                                        sheet.cell(
                                            row=row_num, column=5).value = f'https://pubmed.ncbi.nlm.nih.gov/{pm_id}'  # Pubmed(url)

                                    pmc_id_mo = re.compile(
                                        r'PMCid:PMC(\d+)').search(pmid_record)
                                    if pmc_id_mo:
                                        pmc_id = pmc_id_mo.group(1)
                                        sheet.cell(
                                            row=row_num, column=6).value = f'https://www.ncbi.nlm.nih.gov/pmc/{pmc_id}'  # Pubmed Centrial(url)

        now = datetime.datetime.now()
        wb.save(excel_path / f"참고문헌_일괄등록_{now.strftime('%Y%m%d_%H%M%S')}.xlsx")

    except Exception as e:
        raise Exception
